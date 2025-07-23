"""
封装的需求：
    数据读取：
        封装一个可以读取任意Excel文件的方法，可以指定读取的表单
    数据写入：
        文件名：
        表单名：
        行:
        列：
        写入内容：

"""
import openpyxl
class HandleExcel:
    def __init__(self, file_name, sheet_name):
        """

        :param file_name: 文件名
        :param sheet_name: 表单名
        """
        self.file_name = file_name
        self.sheet_name = sheet_name

    # 未修改版本
    # def read_data(self, file_name, sheet_name):
    #     # 数据读取
    #     workbook = openpyxl.load_workbook(file_name)
    #     sh = workbook[sheet_name]
    #     # 获取所有行
    #     res = list(sh.rows)
    #     # 获取excel中的第一行数据
    #     title = [i.value for i in res[0]]
    #     case = []
    #     # 遍历除第一行外的所有行的数据
    #     for item in res[1:]:
    #         # 获取其他行的数据
    #         data = [i.value for i in item]
    #         # 第一行的数据和当前的数据打包为字典
    #         dic = dict(zip(title, data))
    #         # 把字典添加到列表中
    #         case.append(dic)
    #     return case
    # 修改版本
    def read_data(self):
        """
        :return: 读取完成的数据回传
        """
        # 数据读取
        workbook = openpyxl.load_workbook(self.file_name)
        sh = workbook[self.sheet_name]
        # 获取所有行
        res = list(sh.rows)
        # 获取excel中的第一行数据
        title = [i.value for i in res[0]]
        case = []
        # 遍历除第一行外的所有行的数据
        for item in res[1:]:
            # 获取其他行的数据
            data = [i.value for i in item]
            # 第一行的数据和当前的数据打包为字典
            dic = dict(zip(title, data))
            # 把字典添加到列表中
            case.append(dic)
        return case


    def write_data(self, row, column, value):
        """
        :param row: 写入的行
        :param column: 写入的列
        :param value: 需要写入的值
        :return: 无
        """
        # 数据写入前读取文件
        # 加载工作簿对象
        workbook = openpyxl.load_workbook(self.file_name)
        sh = workbook[self.sheet_name]
        # 数据写入
        sh.cell(row=row, column=column, value=value)
        # 数据保存
        workbook.save(self.file_name)


    # 未改进版本读取数据过程
    # excel = HandleExcel()
    # res = excel.read_data(r"/Users/scan/Desktop/柠檬班/13/testwork_02/testdata.xlsx", 'test')
    # print(res)

    # 改进版本读取数据过程
excel = HandleExcel(r"/Users/scan/Desktop/柠檬班/13/demo_03/data.xlsx", 'Sheet1')
res = excel.write_data(1,1,"tttttt")
print(res)

    