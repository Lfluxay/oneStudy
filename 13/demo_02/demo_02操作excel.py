"""

"""
import openpyxl

# ---------------基本操作-------------
# 1.加载excel文件为工作簿对象
# workbook = openpyxl.load_workbook("test001.xlsx")
# 获取所有的表单名
# print(workbook.sheetnames)

# 2.选中表单
# sh = workbook["Sheet1"]
# print(sh)

# 3.读取数据
# c = sh.cell(row=1, column=1)
# print(c.value)


# ----------------读取全部数据-----------------------
# workbook = openpyxl.load_workbook("test001.xlsx")
# sh = workbook["Sheet1"]

# rows：按行获取表单中所有的格子,每一行的格子放到一个元组中
# res = list(sh.rows)
# print(res)

# 按列读取
# res1 = list(sh.columns)
# print(res1)


# ----------------读取数据存为列表嵌套字典-----------------------
workbook = openpyxl.load_workbook("test001.xlsx")
sh = workbook["test"]

res = list(sh.rows)
# 写成一行
# title = res[0]
# print(title)
# result = []
# for i in title:
#     print(i.value)
title = [i.value for i in res[0]]
print(title)

# 对此进行改进--------
# r1 = [i.value for i in res[1]]
# print(r1)
# print(dict(zip(title,r1)))
#
# r2 = [i.value for i in res[2]]
# print(r2)
# print(dict(zip(title,r2)))
#
# r3 = [i.value for i in res[3]]
# print(r2)
# print(dict(zip(title,r3)))

# 改进方法---------
case = []
for item in res[1:]:
    data = [i.value for i in item]
    dic = dict(zip(title, data))
    case.append(dic)
print(case)